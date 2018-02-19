
from flask import Flask, request, render_template, jsonify
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import json
from tinydb import TinyDB, Query

MAX_FILE_SIZE = 1024 * 1024 * 16 + 1
ALLOWED_EXTENSIONS = set(['xlsx'])
MODE_TABLE = '1'
MODE_GRAPH = '0'
DB_PATH = 'db/db.json'
DOWNLOADS_PATH = 'downloads/'

BLUE_CONSOLE_COLOR = '\033[94m'
END_CONSOLE_COLOR = '\033[0m'

app = Flask(__name__)


# this function here for Python 3.4 and lower versions
def merge_two_dicts(firstDict, secondDict):
    result = firstDict.copy()
    result.update(secondDict)
    return result


# Check if uploaded file is allowed
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS


# Saving new info from donwloaded Excel file to DB
def save_to_db(filename):
    data = read_file(filename)

    print()
    print(BLUE_CONSOLE_COLOR + 'Saving new data to DB...' + END_CONSOLE_COLOR)

    db = TinyDB(DB_PATH)
    qry = Query()

    # print(db.search(qry.id >= 0))
    info = {'id': len(db), 'File': filename, 'Upload date': 14, 'Upload time': 15}
    newEntry = merge_two_dicts(info, data)
    db.insert(newEntry)
    print('New saved DB entry: ')
    print(db.search(qry.id == (len(db) - 1)))
    print(BLUE_CONSOLE_COLOR + 'Saving successful' + END_CONSOLE_COLOR)
    print()

    return True


# Reading data from DB
# MODE_TABLE - put in reply data only for table update
# MODE_GRAPH - put in reply all data for graph rendering
def read_from_db(mode, entryid):
    print()
    print(BLUE_CONSOLE_COLOR + 'Reading from DB...' + END_CONSOLE_COLOR)

    db = TinyDB(DB_PATH)
    qry = Query()

    if mode == MODE_TABLE:
        dbData = db.search(qry.id >= 0)
        reply = {"columns": ["#", "File", "Upload date", "Upload time"], "rows": []}
        print('Forming JSON with table data:')
        for entry in dbData:
            row = {key: entry[key] for key in ['id', 'File', 'Upload date', 'Upload time']}
            reply['rows'].append(row)

    elif mode == MODE_GRAPH:
        print('Forming JSON graph data:')
        dbData = db.search(qry.id == entryid)
        reply = {"nodes": [], "links": []}
        for entry in dbData:
            reply.update(entry)

    else:
        reply = {}

    print(reply)
    print()
    return reply


# Reading just uploaded Excel file
def read_file(name):
    print()
    print(BLUE_CONSOLE_COLOR + 'Reading Excel file...' + END_CONSOLE_COLOR)
    wb = load_workbook(name)
    sheet = wb['Лист1']
    data = {}
    data['nodes'] = []
    data['links'] = []
    unique = set()
    flag = True
    i = 2
    print('Unique nods:')
    while flag == True:
        i += 1
        value_col_3 = sheet.cell(row=i, column=3).value
        if value_col_3:
            if (value_col_3 not in unique):
                data['nodes'].append({'id': str(value_col_3)})
                unique.add(value_col_3)
                print(value_col_3)
        else:
            flag = False

    print('Links:')
    flag = True
    i = 2
    while flag:
        i += 1
        value_col_1 = sheet.cell(row=i, column=1).value
        value_col_2 = sheet.cell(row=i, column=2).value
        if value_col_1:
            if (value_col_1 in unique) and (value_col_2 in unique):
                data['links'].append({'source': str(value_col_1), 'target': str(value_col_2)})
                print(str(value_col_1) + " - " + str(value_col_2))
        else:
            flag = False

    print(BLUE_CONSOLE_COLOR + 'Reading Successful' + END_CONSOLE_COLOR)
    print()
    return data


# Render main page
@app.route('/', methods=["GET"])
def index():
    return render_template("index.html")


# Request for graph json
@app.route('/downloadgraph', methods=["POST"])
def download():
    print()
    print('----')
    r = request.get_json()
    # s = json.loads(r)
    print(r['rownumber'])
    print(BLUE_CONSOLE_COLOR + 'Request for JSON with graph data' + END_CONSOLE_COLOR)
    reply = read_from_db(MODE_GRAPH, r['rownumber'])
    print('JSON was formed: ')
    print(reply)
    print(BLUE_CONSOLE_COLOR + 'Sending JSON with graph info back...' + END_CONSOLE_COLOR)
    print('----')
    print()
    return jsonify(reply)


# Request for table data
@app.route('/updatetable', methods=["GET"])
def update():
    print()
    print('----')
    print(BLUE_CONSOLE_COLOR + 'Request for JSON with table data' + END_CONSOLE_COLOR)
    reply = read_from_db(MODE_TABLE, 0)
    print('JSON was formed.')
    print(BLUE_CONSOLE_COLOR + 'Sending JSON with table info back...' + END_CONSOLE_COLOR)
    print('----')
    print()
    return jsonify(reply)


# Upload new Excel file on server
@app.route('/uploadfile', methods=["POST"])
def upload():
    print()
    print('----')
    print(BLUE_CONSOLE_COLOR + 'User sent new Excel file' + END_CONSOLE_COLOR)
    file = request.files["file"]
    if file and allowed_file(file.filename):
        # file_bytes = file.read(MAX_FILE_SIZE)
        # args["file_size_error"] = len(file_bytes) == MAX_FILE_SIZE
        print('Uploaded new Excel file')
        filename = secure_filename(file.filename)
        file.save(DOWNLOADS_PATH + filename)
        save_to_db(DOWNLOADS_PATH + filename)
        print(BLUE_CONSOLE_COLOR + 'Request was successfully processed, replying 204' + END_CONSOLE_COLOR)
        print('----')
        print()
        return '', 204
    else:
        print(BLUE_CONSOLE_COLOR + 'File has forbidden format, replying 403' + END_CONSOLE_COLOR)
        print('----')
        print()
        return '', 403


# Run server
if __name__ == '__main__':
    app.run(debug=True)
