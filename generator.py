from openpyxl import load_workbook
from random import randrange

wb = load_workbook('static/Test Template.xlsx')
ws = wb.active
sheet = wb['Лист1']

for i in range(3, 104):
    sheet.cell(row=i, column=1).value = randrange(1, 100, 1)
    sheet.cell(row=i, column=2).value = randrange(1, 100, 1)
    sheet.cell(row=i, column=3).value = randrange(1,100,1)



wb.save('test examples/Test' + str(randrange(1,10000000,1)) + '.xlsx')